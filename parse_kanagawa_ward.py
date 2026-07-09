#!/usr/bin/env python3
"""
REINS_Kanagawa 区単位 成約データ(Excel) -> reins_kanagawa_data.json

各区は3ファイル(マンション/戸建/土地)で構成される想定:
    港北区マンション成約データ.xlsx
    港北区戸建成約データ.xlsx
    港北区土地成約データ.xlsx

Usage:
    python3 parse_kanagawa_ward.py --ward 港北区 \
        --mansion 港北区マンション成約データ.xlsx \
        --house 港北区戸建成約データ.xlsx \
        --land 港北区土地成約データ.xlsx \
        --out reins_kanagawa_data.json

複数区がある場合は既存のJSONに追記していく(--out に既存ファイルを指定すればマージされる)。
"""
import argparse
import json
import sys
from pathlib import Path

import openpyxl


def read_table_sheet(path, sheet_name, drop_total_row=True):
    """汎用: タイトル行+ヘッダー行+データ行、という構造のシートを
    {headers: [...], rows: [[...], ...]} として読む。
    列構成はカテゴリ(マンション/戸建/土地)ごとに違うため、ヘッダーをそのまま使う。"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    rows_iter = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows_iter:
        return None
    headers = [h for h in rows_iter[0] if h is not None]
    data_rows = []
    for r in rows_iter[1:]:
        if r[0] is None:
            continue
        if drop_total_row and str(r[0]).startswith("合計"):
            continue
        vals = list(r[: len(headers)])
        data_rows.append(vals)
    return {"headers": headers, "rows": data_rows}


def read_overview(path):
    """データ概要シートから期間・件数を、成約データ一覧から統計を出す"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["成約データ一覧"]
    rows = list(ws.iter_rows(min_row=3, values_only=True))  # skip title+header
    headers = [c for c in ws.iter_rows(min_row=2, max_row=2, values_only=True)][0]
    header_idx = {h: i for i, h in enumerate(headers) if h}

    price_i = header_idx.get("価格(万円)")
    sqm_i = header_idx.get("㎡単価(万円)")
    date_i = header_idx.get("成約年月日")

    prices = [r[price_i] for r in rows if price_i is not None and r[price_i] is not None]
    sqm_prices = [r[sqm_i] for r in rows if sqm_i is not None and r[sqm_i] is not None]
    dates = [r[date_i] for r in rows if date_i is not None and r[date_i] is not None]

    count = len(prices)
    avg_price = round(sum(prices) / count) if count else None
    avg_sqm = round(sum(sqm_prices) / len(sqm_prices), 1) if sqm_prices else None
    period = None
    if dates:
        try:
            period = f"{min(dates)}〜{max(dates)}"
        except TypeError:
            period = None

    return {"count": count, "avgPrice": avg_price, "avgSqmPrice": avg_sqm, "period": period}


def read_price_bands(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["価格帯分布"]
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    bands = []
    for r in rows:
        if not r or r[0] is None:
            continue
        band, count, pct = r[0], r[1], r[-1]
        if count is None:
            continue
        bands.append({"band": str(band), "count": int(count), "pct": float(pct) if pct is not None else None})
    return bands


def read_station_ranking(path, limit=10):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "最寄駅別集計" not in wb.sheetnames:
        return []
    ws = wb["最寄駅別集計"]
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    out = []
    for r in rows[:limit]:
        if not r or r[1] is None:
            continue
        out.append({
            "station": r[1], "count": int(r[2]) if r[2] is not None else None,
            "avgPrice": r[3], "avgSqmPrice": r[5] if len(r) > 5 else None,
        })
    return out


def parse_category(path):
    age_or_tsubo = read_table_sheet(path, "築年代別分布")
    tsubo = read_table_sheet(path, "坪単価帯別分布")
    out = {
        "overview": read_overview(path),
        "priceBands": read_price_bands(path),
        "stationRanking": read_station_ranking(path),
        "walkBands": read_table_sheet(path, "徒歩圏別相場", drop_total_row=True),
    }
    if age_or_tsubo:
        out["ageBands"] = age_or_tsubo
    if tsubo:
        out["tsuboBands"] = tsubo
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--ward", required=True)
    ap.add_argument("--mansion")
    ap.add_argument("--house")
    ap.add_argument("--land")
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    ward_data = {}
    if args.mansion:
        ward_data["中古マンション"] = parse_category(args.mansion)
    if args.house:
        ward_data["戸建"] = parse_category(args.house)
    if args.land:
        ward_data["土地"] = parse_category(args.land)

    out_path = Path(args.out)
    existing = {}
    if out_path.exists():
        existing = json.loads(out_path.read_text(encoding="utf-8"))
    existing.setdefault("wards", {})
    existing["wards"][args.ward] = ward_data
    out_path.write_text(json.dumps(existing, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"→ {args.out} に {args.ward} を書き込みました")
    for cat, d in ward_data.items():
        print(f"  {cat}: {d['overview']['count']}件, 平均{d['overview']['avgPrice']}万円")


if __name__ == "__main__":
    main()
